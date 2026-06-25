import openpyxl
import csv
import os

EXCEL_PATH = r"C:\Users\Admin\Downloads\KIRO PROJETOS\loja de roupas\californiastore-saas/California (LOJA ROUPAS).xlsx"
OUTPUT_DIR = r"C:\Users\Admin\Downloads\KIRO PROJETOS\loja de roupas\californiastore-saas\imports"

# ========== Mapeamentos ==========

GENEROS_MAP = {
    "MASCULINO": "MASCULINO",
    "MASCULINA": "MASCULINO",
    "MASCILINA": "MASCULINO",
    "FEMININO": "FEMININO",
    "FEMININA": "FEMININO",
    "FEMINIA": "FEMININO",
    "UNISSEX": "UNISSEX",
    "UNISEX": "UNISSEX",
    "COMPARTILHAVEL": "UNISSEX",
}

TAMANHOS_MAP = {
    "UNICO": "UN",
    "UN": "UN",
    "GRANDE": "G",
    "PEQUENO": "P",
    "G1": "G",
    "G2": "G",
    "G3": "G",
    "GGG": "GG",
    "XGG": "XXG",
    "EXG": "XXG",
    "2XG": "XXL",
}

CORES_MAP = {
    "BRANCA": "BRANCO",
    "PRETA": "PRETO",
    "AZUL ESCURA": "AZUL ESCURO",
}

CATEGORIAS_MERGE = {
    "JAQUETA HERITAGE": "JAQUETA",
    "JAQUETA ECLIPSE": "JAQUETA",
    "FIVELA LEN": "FIVELA",
}

# ========== Funcoes ==========

def clean_genero(val):
    if not val:
        return None
    raw = str(val).strip().upper()
    return GENEROS_MAP.get(raw, raw)

def clean_tamanho(val):
    if not val:
        return None
    raw = str(val).strip().upper()
    mapped = TAMANHOS_MAP.get(raw, raw)
    return mapped

def clean_cor(val):
    if not val:
        return None
    raw = str(val).strip().upper()
    return CORES_MAP.get(raw, raw)

def clean_categoria(val):
    if not val:
        return None
    raw = str(val).strip().upper()
    return CATEGORIAS_MERGE.get(raw, raw)

def clean_marca(val):
    if not val:
        return None
    return str(val).strip().upper()

def clean_text(val):
    if not val:
        return None
    return str(val).strip()

def clean_foto_url(val):
    if not val:
        return ""
    url = str(val).strip()
    # Converter Google Drive link para imagem direta
    import re
    match = re.search(r"/file/d/([^/]+)", url)
    if match:
        return f"https://drive.google.com/uc?export=view&id={match.group(1)}"
    return url

# ========== Leitura ==========

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Geral"]

rows_clean = []
errors = []

# Headers (row 2)
headers_raw = []
for col in range(1, 13):
    headers_raw.append(str(ws.cell(2, col).value or ""))

for row_idx in range(3, ws.max_row + 1):
    cod_interno = ws.cell(row_idx, 1).value
    if cod_interno is None:
        continue

    cod_forn = ws.cell(row_idx, 2).value
    categoria = clean_categoria(ws.cell(row_idx, 3).value)
    marca = clean_marca(ws.cell(row_idx, 4).value)
    descricao = clean_text(ws.cell(row_idx, 5).value)
    genero = clean_genero(ws.cell(row_idx, 6).value)
    tamanho = clean_tamanho(ws.cell(row_idx, 7).value)
    cor = clean_cor(ws.cell(row_idx, 8).value)
    quantidade = ws.cell(row_idx, 9).value
    custo = ws.cell(row_idx, 10).value
    venda = ws.cell(row_idx, 11).value
    fotos = ws.cell(row_idx, 12).value

    # Para perfumes: descricao esta na coluna MARCA (4), nao na coluna 5
    if not descricao and categoria and "PERFUME" in categoria.upper():
        descricao = marca
        marca = ""

    if not descricao:
        errors.append(f"Linha {row_idx}: sem descricao")
        continue
    if not genero:
        errors.append(f"Linha {row_idx}: genero invalido '{ws.cell(row_idx, 6).value}'")
        continue

    try:
        qtd = int(quantidade) if quantidade else 0
    except (ValueError, TypeError):
        qtd = 0
        errors.append(f"Linha {row_idx}: quantidade invalida '{quantidade}'")

    try:
        preco_custo = float(custo) if custo else 0
    except (ValueError, TypeError):
        preco_custo = 0
        errors.append(f"Linha {row_idx}: custo invalido '{custo}'")

    try:
        preco_venda = float(venda) if venda else 0
    except (ValueError, TypeError):
        preco_venda = 0
        errors.append(f"Linha {row_idx}: venda invalida '{venda}'")

    rows_clean.append({
        "codigo_interno": str(cod_interno).strip(),
        "codigo_fornecedor": str(cod_forn).strip() if cod_forn else "",
        "categoria": categoria or "",
        "marca": marca or "",
        "descricao": descricao,
        "genero": genero,
        "tamanho": tamanho or "",
        "cor": cor or "",
        "quantidade": qtd,
        "preco_custo": preco_custo,
        "preco_venda": preco_venda,
        "foto_url": clean_foto_url(fotos),
    })

# ========== Estatisticas ==========

cats = {}
gens = {}
sizes = {}
colors = {}
brands = {}

for r in rows_clean:
    cats[r["categoria"]] = cats.get(r["categoria"], 0) + 1
    gens[r["genero"]] = gens.get(r["genero"], 0) + 1
    if r["tamanho"]:
        sizes[r["tamanho"]] = sizes.get(r["tamanho"], 0) + 1
    if r["cor"]:
        colors[r["cor"]] = colors.get(r["cor"], 0) + 1
    if r["marca"]:
        brands[r["marca"]] = brands.get(r["marca"], 0) + 1

# Agrupar por codigo_interno (variantes)
produtos = {}
for r in rows_clean:
    cod = r["codigo_interno"]
    if cod not in produtos:
        produtos[cod] = {
            "codigo_interno": cod,
            "codigo_fornecedor": r["codigo_fornecedor"],
            "categoria": r["categoria"],
            "marca": r["marca"],
            "descricao": r["descricao"],
            "genero": r["genero"],
            "preco_custo": r["preco_custo"],
            "preco_venda": r["preco_venda"],
            "variantes": [],
        }
    produtos[cod]["variantes"].append({
        "tamanho": r["tamanho"],
        "cor": r["cor"],
        "quantidade": r["quantidade"],
    })

# ========== Output ==========

os.makedirs(OUTPUT_DIR, exist_ok=True)

# CSV de todas as linhas limpas
csv_path = os.path.join(OUTPUT_DIR, "produtos_limpo.csv")
fieldnames = ["codigo_interno", "codigo_fornecedor", "categoria", "marca", "descricao", "genero", "tamanho", "cor", "quantidade", "preco_custo", "preco_venda", "foto_url"]

with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows_clean)

# CSV de produtos unicos (agrupados)
produtos_path = os.path.join(OUTPUT_DIR, "produtos_agrupados.csv")
with open(produtos_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["codigo_interno", "codigo_fornecedor", "categoria", "marca", "descricao", "genero", "preco_custo", "preco_venda", "qtd_variantes", "estoque_total"])
    writer.writeheader()
    for p in produtos.values():
        estoque = sum(v["quantidade"] for v in p["variantes"])
        writer.writerow({
            "codigo_interno": p["codigo_interno"],
            "codigo_fornecedor": p["codigo_fornecedor"],
            "categoria": p["categoria"],
            "marca": p["marca"],
            "descricao": p["descricao"],
            "genero": p["genero"],
            "preco_custo": p["preco_custo"],
            "preco_venda": p["preco_venda"],
            "qtd_variantes": len(p["variantes"]),
            "estoque_total": estoque,
        })

# ========== Relatorio ==========

print("=" * 60)
print("RELATORIO DE LIMPEZA")
print("=" * 60)
print(f"Linhas originais: 640")
print(f"Linhas limpas: {len(rows_clean)}")
print(f"Erros/ignoradas: {len(errors)}")
print(f"Produtos unicos: {len(produtos)}")
print()
print(f"CSV limpo: {csv_path}")
print(f"CSV agrupado: {produtos_path}")
print()
print(f"=== CATEGORIAS ({len(cats)}) ===")
for k, v in sorted(cats.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
print()
print(f"=== GENEROS ({len(gens)}) ===")
for k, v in sorted(gens.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
print()
print(f"=== MARCAS ({len(brands)}) ===")
for k, v in sorted(brands.items(), key=lambda x: -x[1])[:15]:
    print(f"  {k}: {v}")
print(f"  ... e mais {len(brands) - 15}")
print()
print(f"=== ERROS ({len(errors)}) ===")
for e in errors[:10]:
    print(f"  {e}")
if len(errors) > 10:
    print(f"  ... e mais {len(errors) - 10}")
